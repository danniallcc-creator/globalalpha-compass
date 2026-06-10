#!/usr/bin/env python3
"""
Update HS codes in all category JSON files and category_index.json
using data from the HS code Excel file.

Data sources:
  - Sheet1 ("Sheet1-建材行业hs编码"): Building material specific HS codes
    Columns: 一级类目, 二级, 海关hs编码
  - Sheet2 ("Sheet2-hs编码大全"): Comprehensive HS code mapping
    Columns: hscode6位, hscode6位描述, 已映射的一级行业, 已映射的二级行业

Logic:
  - For building materials (建材与房地产, 工程及建材机械):
    Use Sheet1 data first (more specific), supplement with Sheet2 data.
  - For all other categories:
    Use Sheet2 data matched by 已映射的一级行业 and 已映射的二级行业.
"""

import json
import os
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip3 install openpyxl")
    sys.exit(1)

# Paths
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.expanduser("~/Desktop/hs编码大全.xlsx")
DATA_DIR = os.path.join(BASE_DIR, "data")
CATEGORY_INDEX_PATH = os.path.join(DATA_DIR, "category_index.json")
CATEGORIES_DIR = os.path.join(DATA_DIR, "categories")

# L1 categories that should use Sheet1 (building materials)
BUILDING_MATERIAL_L1 = {"建材与房地产", "工程及建材机械"}


def hs_code_to_str(code):
    """Convert a numeric HS code to a 6-digit string."""
    s = str(int(code))
    if len(s) >= 6:
        return s[:6]
    # Pad with trailing zeros for codes shorter than 6 digits
    return s.ljust(6, "0")


def derive_prefix(hs_codes_list):
    """Derive the hs_code_prefix from a list of HS code strings.

    Strategy: find the most common 4-digit prefix among the codes,
    then format as '{prefix}xx'. If no clear majority, use the prefix
    of the first code.
    """
    if not hs_codes_list:
        return "99xx"
    prefix_counts = defaultdict(int)
    for code in hs_codes_list:
        if len(code) >= 4:
            prefix_counts[code[:4]] += 1
    if prefix_counts:
        most_common = max(prefix_counts, key=prefix_counts.get)
        return most_common[:2] + "xx"
    return hs_codes_list[0][:2] + "xx"


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: Excel file not found: {EXCEL_PATH}")
        sys.exit(1)

    print(f"Reading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    sheet_names = wb.sheetnames
    print(f"  Sheets: {sheet_names}")

    # ===== Read Sheet1: Building materials HS codes =====
    ws1 = wb[sheet_names[0]]
    # Mapping: 二级 name -> list of HS code strings
    sheet1_by_l2 = defaultdict(list)
    for row in ws1.iter_rows(min_row=2, values_only=True):
        l2_name = row[1]
        hs_raw = row[2]
        if l2_name is not None and hs_raw is not None:
            sheet1_by_l2[str(l2_name).strip()].append(hs_code_to_str(hs_raw))

    sheet1_count = sum(len(v) for v in sheet1_by_l2.values())
    print(f"  Sheet1: {len(sheet1_by_l2)} L2 categories, {sheet1_count} HS code entries")

    # ===== Read Sheet2: Full HS code mapping =====
    ws2 = wb[sheet_names[1]]
    # Mapping: (一级行业, 二级行业) -> list of HS code strings
    sheet2_by_l1l2 = defaultdict(list)
    for row in ws2.iter_rows(min_row=2, values_only=True):
        hs_raw = row[0]
        l1_ind = row[2]
        l2_ind = row[3]
        if (
            hs_raw is not None
            and l1_ind is not None
            and l1_ind != "-"
            and l2_ind is not None
            and l2_ind != "-"
        ):
            key = (str(l1_ind).strip(), str(l2_ind).strip())
            sheet2_by_l1l2[key].append(hs_code_to_str(hs_raw))

    sheet2_count = sum(len(v) for v in sheet2_by_l1l2.values())
    print(f"  Sheet2: {len(sheet2_by_l1l2)} L1-L2 pairs, {sheet2_count} HS code entries")

    wb.close()

    # ===== Load category index =====
    print(f"\nLoading category index: {CATEGORY_INDEX_PATH}")
    with open(CATEGORY_INDEX_PATH, "r", encoding="utf-8") as f:
        index_data = json.load(f)

    total = len(index_data["categories"])
    print(f"  {total} categories in index")

    # ===== Process each category =====
    stats = {
        "updated": 0,
        "no_match": 0,
        "skipped_no_file": 0,
        "kept_existing": 0,
    }
    unmatched_categories = []

    print("\n" + "=" * 70)
    print("Processing categories...")
    print("=" * 70)

    for cat in index_data["categories"]:
        l1_cn = cat["l1_cn"]
        name_cn = cat["name_cn"]
        l1_slug = cat["l1_slug"]
        l2_slug = cat["l2_slug"]

        json_path = os.path.join(CATEGORIES_DIR, l1_slug, f"{l2_slug}.json")

        # Determine which data source to use
        is_building = l1_cn in BUILDING_MATERIAL_L1

        new_hs_codes = set()

        if is_building:
            # Building materials: Sheet1 first, then Sheet2 as supplement
            if name_cn in sheet1_by_l2:
                new_hs_codes.update(sheet1_by_l2[name_cn])

            sheet2_key = (l1_cn, name_cn)
            if sheet2_key in sheet2_by_l1l2:
                new_hs_codes.update(sheet2_by_l1l2[sheet2_key])

        else:
            # Non-building: Sheet2 only
            sheet2_key = (l1_cn, name_cn)
            if sheet2_key in sheet2_by_l1l2:
                new_hs_codes.update(sheet2_by_l1l2[sheet2_key])

        # Convert to sorted list
        new_hs_codes_list = sorted(new_hs_codes)

        if not new_hs_codes_list:
            # No HS codes found from Excel - keep existing values
            stats["no_match"] += 1
            unmatched_categories.append(f"  {l1_cn} / {name_cn} (slug: {l2_slug})")
            continue

        # Derive the new prefix
        new_prefix = derive_prefix(new_hs_codes_list)

        # Update category_index entry
        old_prefix = cat.get("hs_code_prefix", "N/A")
        cat["hs_code_prefix"] = new_prefix

        # Update category JSON file
        if not os.path.exists(json_path):
            stats["skipped_no_file"] += 1
            print(f"  [SKIP-NO-FILE] {l1_cn}/{name_cn} -> {json_path}")
            continue

        with open(json_path, "r", encoding="utf-8") as f:
            cat_data = json.load(f)

        old_codes = cat_data.get("hs_codes", [])
        cat_data["hs_codes"] = new_hs_codes_list

        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(cat_data, f, ensure_ascii=False, indent=2)

        stats["updated"] += 1
        source_label = "Sheet1" if is_building and name_cn in sheet1_by_l2 else "Sheet2"
        print(
            f"  [OK] {l1_cn}/{name_cn}: "
            f"{old_codes} -> {new_hs_codes_list} "
            f"(prefix: {old_prefix}->{new_prefix}, src: {source_label})"
        )

    # ===== Save updated category index =====
    print(f"\nSaving updated category index...")
    with open(CATEGORY_INDEX_PATH, "w", encoding="utf-8") as f:
        json.dump(index_data, f, ensure_ascii=False, indent=2)
    print(f"  Saved: {CATEGORY_INDEX_PATH}")

    # ===== Print summary =====
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"  Total categories in index:    {total}")
    print(f"  Updated (JSON + index):       {stats['updated']}")
    print(f"  No HS code match in Excel:    {stats['no_match']}")
    print(f"  Skipped (no JSON file):       {stats['skipped_no_file']}")

    if unmatched_categories:
        print(f"\n  Categories with NO match in Excel ({len(unmatched_categories)}):")
        for uc in unmatched_categories:
            print(f"    {uc}")

    # ===== Verify a few samples =====
    print("\n" + "=" * 70)
    print("VERIFICATION (sample entries)")
    print("=" * 70)
    samples = [
        ("建材与房地产", "prefabricated-buildings"),
        ("建材与房地产", "building-boards"),
        ("建材与房地产", "stone"),
        ("个人护理及家庭清洁", "hygiene-products"),
        ("家居园艺", "home-decor"),
        ("五金工具", "hand-tools"),
        ("工程及建材机械", "engineering-construction-machinery"),
    ]
    for l1, l2 in samples:
        jpath = os.path.join(CATEGORIES_DIR, l1, f"{l2}.json")
        if os.path.exists(jpath):
            with open(jpath, "r", encoding="utf-8") as f:
                d = json.load(f)
            print(f"  {l1}/{l2}: hs_codes={d.get('hs_codes', 'N/A')}")
        else:
            print(f"  {l1}/{l2}: (file not found)")

    print("\nDone!")


if __name__ == "__main__":
    main()
